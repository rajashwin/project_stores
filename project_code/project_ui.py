##--gui module
import time
import datetime

from tkinter import *
import tkinter as tk
from tkinter import font
import tkinter.messagebox
import tkentrycomplete

from PIL import Image, ImageTk
import sqlite3
#from openpyxl import openpyxl.load_workbook
import openpyxl
from pathlib import Path
import os
data_folder = "/home/king/PycharmProjects/project_stores"
# # make database and users (if not exists already) table at programme start up
# with sqlite3.connect('quit1.db') as d \b:
# 	c = db.cursor()
#
#
# #c.execute('CREATE TABLE IF NOT EXISTS user (username TEXT PRIMARY KEY  NOT NULL,NAME   TEXT    NOT NULL,EMPID   INT   NOT NULL,PASSWORD  TEXT    NOT NULL,);')
#
# c.execute('CREATE TABLE IF NOT EXISTS user (username TEXT NOT NULL PRIMARY KEY,'
# 		  'password TEXT NOT NULL,'
# 		  'NAME   TEXT    NOT NULL,'
# 		  'EMPID   INT   NOT NULL);')
# db.commit()
# db.close()

class rpi_UI:
	def __init__(self):
		if __name__ == '__main__':
			self.window = tk.Tk()
			self.window.overrideredirect(False)
			# ===============================================================
			##--background support for all UI
			self.window.title("Penguin Version")
			self.window.resizable(width=900, height=600)

			self.GUIFrame_default = tk.Frame(self.window, width=900, height=600)
			self.GUIFrame_default.pack(expand=False, anchor=CENTER)
			self.GUIFrame_default.pack()

			self.adminLogin=0
			self.loginName=""
			# ===============================================================
			##--Supporting IMAGES
			
			self.photo_user_icon = tk.PhotoImage(file=str(data_folder)+"/Images/icon/icons8-user.png")
			self.photo_user_pwd = tk.PhotoImage(file=str(data_folder)+"/Images/icon/icons8-password.png")
			self.photo_wrong = tk.PhotoImage(file=str(data_folder)+"/Images/icon/wrong-icon5.gif")
			self.photo_correct = tk.PhotoImage(file=str(data_folder)+"/Images/icon/correct.gif")

			self.photo_stockLogo = tk.PhotoImage(file=str(data_folder)+"/Images/banners/stock.png")
			self.photo_westtek = tk.PhotoImage(file=str(data_folder)+"/Images/banners/westtek.png")

			self.photo_asiLogo = tk.PhotoImage(file=str(data_folder)+"/Images/banners/asi_logo_1.png")
			self.photo_user_scn = tk.PhotoImage(file=str(data_folder)+"/Images/icon/icons8-barcode-scanner.png")

			self.update_log=tk.PhotoImage(file=str(data_folder)+"/Images/icon/update_icon1.png")

			self.wb = openpyxl.load_workbook(str(data_folder)+"/data_storage/store_details.xlsx")
			self.sheet = self.wb["Sheet1"]
			self.row_count = self.sheet.max_row
			self.column_count = self.sheet.max_column


			##--FONT TYPE
			self.helv36_login = font.Font(family='Helvetica', size=16, weight='bold')
			self.helv8_login = font.Font(family='Helvetica', size=8, weight='normal')
			self.times_16 = font.Font(family="Helvetica", size=16, weight='normal')
			self.page_one()
			self.window.mainloop()


	def page_one(self):
		##--Login page
		self.GUIFrame_login = tk.Frame(self.GUIFrame_default, width=900, height=600)
		self.GUIFrame_login.pack(expand=False, anchor=CENTER)
		self.GUIFrame_login.place(x=0, y=0)

		##--Login widiget
		self.login_label = tk.Label(self.GUIFrame_login, font=self.helv36_login, text="Login ID", fg='blue',)
		self.login_label.place(x=500, y=350)

		self.loginid = StringVar()
		self.login_entry = tk.Entry(self.GUIFrame_login, font=self.helv36_login, textvariable=self.loginid, width=16)
		self.login_entry.place(x=650, y=350)
		self.login_entry.configure(bg='snow')

		self.ltext_usr = tk.Label(self.GUIFrame_login, image=self.photo_user_icon)
		self.ltext_usr.place(x=610, y=350)
		self.login_entry.bind("<Button-1>", self.login_entry_event)

		##--password widiget
		self.Password_label = tk.Label(self.GUIFrame_login, font=self.helv36_login, text="Password", fg='blue')
		self.Password_label.place(x=500, y=400)

		self.password = StringVar()
		self.password_entry = tk.Entry(self.GUIFrame_login, font=self.helv36_login, textvariable=self.password,
		                               width=16, show="*")
		self.password_entry.place(x=650, y=400)
		self.password_entry.configure(bg='snow')

		self.ltext_pwd = tk.Label(self.GUIFrame_login, image=self.photo_user_pwd)
		self.ltext_pwd.place(x=610, y=400)
		self.password_entry.bind("<Button-1>", self.password_entry_event)
#---check user image check

		self.usr_img_check = tk.Label(self.GUIFrame_login)
		self.usr_img_check.place(x=860, y=350)
		self.pwd_img_check = tk.Label(self.GUIFrame_login)
		self.pwd_img_check.place(x=860, y=400)


		self.ltext1 = tk.Label(self.GUIFrame_login, image=self.photo_stockLogo)
		self.ltext1.place(x=0, y=100)

		self.ltext2 = tk.Label(self.GUIFrame_login, image=self.photo_westtek)
		self.ltext2.place(x=10, y=10)

		self.ltext3 = tk.Label(self.GUIFrame_login, image=self.photo_asiLogo)
		self.ltext3.place(x=500, y=100)

		self.login_text = tk.Label(self.GUIFrame_login, font=self.helv36_login, text=" USER LOGIN ",borderwidth=1, relief="solid")
		self.login_text.place(x=650, y=480)
		self.login_text.bind("<Button-1>",self.pwd_enter)

		self.ltext_scn = tk.Label(self.GUIFrame_login, image=self.photo_user_scn)
		self.ltext_scn.place(x=550, y=460)

		self.CheckVar = IntVar()
		self.checkButton_reg = Checkbutton(self.GUIFrame_login, text="Admin Register", variable=self.CheckVar,onvalue=1, offvalue=0, height=5, width=20)
		self.checkButton_reg.place(x=700, y=20)

	def login_entry_event(self, key):
		self.password_entry.configure(bg='snow')
		self.login_entry.configure(bg='SeaGreen1')

	def password_entry_event(self, key):
		self.password_entry.configure(bg='SeaGreen1')
		self.login_entry.configure(bg='snow')

	def pwd_enter(self,key):
		if (self.login_entry.get() == "westtek" and self.password_entry.get() == "West@1234"):
			self.loginName ="westtek"
			self.usr_img_check.configure(image=self.photo_correct)
			self.pwd_img_check.configure(image=self.photo_correct)
			if(self.CheckVar.get()==1):
				self.admin_Register()

			else:
				self.adminLogin = 1
				self.GUIFrame_login.destroy()
				self.page_two()


		elif (self.login_entry.get() == "westtek"):
			self.usr_img_check.configure(image=self.photo_correct)
			self.pwd_img_check.configure(image=self.photo_wrong)

		elif (self.password_entry.get() == "West@1234"):
			self.usr_img_check.configure(image=self.photo_wrong)
			self.pwd_img_check.configure(image=self.photo_correct)


		else:
			wb_obj = openpyxl.load_workbook("/home/king/PycharmProjects/project_stores/data_storage/user_data.xlsx")
			sheet_obj = wb_obj.active
			m_row = sheet_obj.max_row
			username = self.login_entry.get()
			password = self.password_entry.get()
			user_check=0;
			# Loop will print all values
			# of first column
			for i in range(1, m_row + 1):
				cell_obj1 = sheet_obj.cell(row=i, column=5)
				cell_obj2 = sheet_obj.cell(row=i, column=6)
				if(username == cell_obj1.value):
					self.usr_img_check.configure(image=self.photo_correct)
					user_check=1

					if( password == cell_obj2.value):
						self.loginName=username
						self.pwd_img_check.configure(image=self.photo_correct)
						user_check=2
						self.adminLogin=0
						break
					else:
						self.pwd_img_check.configure(image=self.photo_wrong)
				elif(user_check==0):
					self.usr_img_check.configure(image=self.photo_wrong)
			if(user_check==2):
				self.GUIFrame_login.destroy()
				self.page_two()

	def admin_Register(self):
		# --Login page
		self.register_page()
		pass

	def register_page(self):
		self.GUIFrame_login.destroy()
		self.GUIFrame_admin = tk.Frame(self.GUIFrame_default, width=900, height=600)
		self.GUIFrame_admin.pack(expand=False, anchor=CENTER)
		self.GUIFrame_admin.place(x=0, y=0)

		self.ltext2 = tk.Label(self.GUIFrame_admin, image=self.photo_westtek)
		self.ltext2.place(x=10, y=10)
###---------NAME
		self.reg_name_label = tk.Label(self.GUIFrame_admin, font=self.helv36_login, text="NAME:", fg='blue')
		self.reg_name_label.place(x=100, y=100)

		self.name_reg = StringVar()
		self.reg_name_entry = tk.Entry(self.GUIFrame_admin, font=self.helv36_login, textvariable=self.name_reg, width=20)
		self.reg_name_entry.place(x=200, y=100)

###---------Employee ID
		self.reg_emp_label = tk.Label(self.GUIFrame_admin, font=self.helv36_login, text="EMP ID:", fg='blue')
		self.reg_emp_label.place(x=100, y=150)

		self.emp_reg = StringVar()
		self.emp_id_entry = tk.Entry(self.GUIFrame_admin, font=self.helv36_login, textvariable=self.emp_reg,
		                               width=20)
		self.emp_id_entry.place(x=200, y=150)

###---------Employee Email-ID
		self.reg_email_label = tk.Label(self.GUIFrame_admin, font=self.helv36_login, text="EMAIL ID:", fg='blue')
		self.reg_email_label.place(x=80, y=200)

		self.email_reg = StringVar()
		self.emp_email_entry = tk.Entry(self.GUIFrame_admin, font=self.helv36_login, textvariable=self.email_reg,
									 width=20)
		self.emp_email_entry.place(x=200, y=200)

###---------USER NAME
		self.reg_usr_label = tk.Label(self.GUIFrame_admin, font=self.helv36_login, text="USER NAME:", fg='blue')
		self.reg_usr_label.place(x=50, y=250)

		self.usr_reg = StringVar()
		self.usr_reg_entry = tk.Entry(self.GUIFrame_admin, font=self.helv36_login, textvariable=self.usr_reg,
		                             width=20)
		self.usr_reg_entry.place(x=200, y=250)

###---------USER password
		self.reg_pwd_label = tk.Label(self.GUIFrame_admin, font=self.helv36_login, text="PASSWORD:", fg='blue')
		self.reg_pwd_label.place(x=50, y=300)

		self.pwd_reg = StringVar()
		self.pwd_reg_entry = tk.Entry(self.GUIFrame_admin, font=self.helv36_login, textvariable=self.pwd_reg,show="*",width=20)
		self.pwd_reg_entry.place(x=200, y=300)


		# ---scrollbar frame view registered name
		self.GUIFramescrollbar_reg = tk.Frame(self.GUIFrame_admin, width=50, height=400)
		self.GUIFramescrollbar_reg.place(x=500, y=100)

		# ---scrollbar initilization
		self.scrollbar_y_reg = Scrollbar(self.GUIFramescrollbar_reg, width=40)
		self.scrollbar_y_reg.pack(side=RIGHT, fill=Y)

		self.scrollbar_x_reg = Scrollbar(self.GUIFramescrollbar_reg, width=32, orient=HORIZONTAL)
		self.scrollbar_x_reg.pack(side=BOTTOM, fill=X)

		# ---textbox frame
		self.txt_reg = tk.Text(self.GUIFramescrollbar_reg, height=20, width=50, wrap=tk.NONE,
						   xscrollcommand=self.scrollbar_x_reg.set,
						   yscrollcommand=self.scrollbar_y_reg.set)
		self.txt_reg.pack()

		# ---adding scrollbar to text
		self.scrollbar_y_reg.config(command=self.txt_reg.yview)
		self.scrollbar_x_reg.config(command=self.txt_reg.xview)

		self.txt_reg.configure(state='normal')
		self.txt_reg.delete(1.0, END)

		# register database
		wb_user = openpyxl.load_workbook(str(data_folder) + "/data_storage/user_data.xlsx")
		sheet_user =wb_user["Sheet1"]
		row_count_user = sheet_user.max_row
		column_count_user = sheet_user.max_column


		for i in range(1, row_count_user + 1):
			if(i==2):
				self.txt_reg.insert(tk.END, "---------------------------------------------------------------------------------------")
				self.txt_reg.insert(tk.INSERT, '\n')
			for j in range(1, column_count_user + 1):
				data = sheet_user.cell(row=i, column=j).value
				if (j == 1):
					self.txt_reg.insert(tk.END, str(data).ljust(5)+' | ')
				elif (j == 2):
					self.txt_reg.insert(tk.END, str(data).ljust(20)+' | ')
				elif(j == 3):
					self.txt_reg.insert(tk.END, str(data).ljust(11)+' | ')
				elif(j == 4):
					self.txt_reg.insert(tk.END, str(data).ljust(40)+' | ')

			self.txt_reg.insert(tk.INSERT,'\n')
		self.txt_reg.configure(state='disable')




###---------save REGISTER BUTTON
		self.Button_reg = tk.Button(self.GUIFrame_admin, font=self.helv36_login, bd=1, text='SAVE ',
		                            command=lambda: self.save_admin_Register(), width=10)
		self.Button_reg.place(x=700, y=520)
###---------Back BUTTON
		self.Button_bck = tk.Button(self.GUIFrame_admin, font=self.helv36_login, bd=1, text='BACK ',
		                            command=lambda: self.back_admin_Register(), width=10)
		self.Button_bck.place(x=100, y=520)

	def save_admin_Register(self):
		if(len(self.name_reg.get()) == 0):
			tk.messagebox.showerror("Error", "Fill Name")
		elif(len(self.emp_reg.get()) == 0):
			tk.messagebox.showerror("Error", "Fill Employee ID")
		elif (len(self.email_reg.get()) == 0):
			tk.messagebox.showerror("Error", "Fill Email-ID")
		elif (len(self.usr_reg.get()) == 0):
			tk.messagebox.showerror("Error", "Fill User name")
		elif (len(self.pwd_reg.get()) == 0):
			tk.messagebox.showerror("Error", "Fill Password")
		else:
			admin_reg_wb = openpyxl.load_workbook("/home/king/PycharmProjects/project_stores/data_storage/user_data.xlsx")
			admin_reg_sheet = admin_reg_wb["Sheet1"]
			admin_regrow = admin_reg_sheet.max_row + 1
			admin_reg_update_value = [admin_reg_sheet.max_row ,self.name_reg.get(),self.emp_reg.get(),self.email_reg.get(),
									  self.usr_reg.get(),self.pwd_reg.get()]
			print(admin_reg_update_value)
			for col, entry in enumerate(admin_reg_update_value, start=1):
				admin_reg_sheet.cell(row=admin_regrow, column=col, value=entry)
			admin_reg_wb.save("/home/king/PycharmProjects/project_stores/data_storage/user_data.xlsx")

			# register database
			self.txt_reg.configure(state='normal')
			self.txt_reg.delete(1.0, END)
			wb_user = openpyxl.load_workbook(str(data_folder) + "/data_storage/user_data.xlsx")
			sheet_user = wb_user["Sheet1"]
			row_count_user = sheet_user.max_row
			column_count_user = sheet_user.max_column

			for i in range(1, row_count_user + 1):
				if (i == 2):
					self.txt_reg.insert(tk.END,
										"---------------------------------------------------------------------------------------")
					self.txt_reg.insert(tk.INSERT, '\n')
				for j in range(1, column_count_user + 1):
					data = sheet_user.cell(row=i, column=j).value
					if (j == 1):
						self.txt_reg.insert(tk.END, str(data).ljust(5) + ' | ')
					elif (j == 2):
						self.txt_reg.insert(tk.END, str(data).ljust(20) + ' | ')
					elif (j == 3):
						self.txt_reg.insert(tk.END, str(data).ljust(11) + ' | ')
					elif (j == 4):
						self.txt_reg.insert(tk.END, str(data).ljust(40) + ' | ')

				self.txt_reg.insert(tk.INSERT, '\n')
			self.GUIFrame_admin.update()
			self.txt_reg.configure(state='disable')

			####---clear All the fields
			self.reg_name_entry.delete(0, END)
			self.emp_id_entry.delete(0, END)
			self.emp_email_entry.delete(0, END)
			self.usr_reg_entry.delete(0, END)
			self.pwd_reg_entry.delete(0, END)

	def back_admin_Register(self):
		self.GUIFrame_admin.destroy()
		self.page_one()

	def page_two(self):
		self.GUIFrame_login.destroy()
		self.GUIFrame_page_two = tk.Frame(self.GUIFrame_default, width=900, height=600)
		self.GUIFrame_page_two.pack(expand=False, anchor=CENTER)
		self.GUIFrame_page_two.place(x=0, y=0)

		self.ltext1 = tk.Label(self.GUIFrame_page_two, image=self.photo_stockLogo)
		self.ltext1.place(x=0, y=100)

		self.ltext2 = tk.Label(self.GUIFrame_page_two, image=self.photo_westtek)
		self.ltext2.place(x=10, y=10)
		###---------Dispense BUTTON
		self.Button_admin_up = tk.Button(self.GUIFrame_page_two, font=self.helv36_login, bd=2, text='ADMIN UPDATE',
		                            command=lambda: self.admin_update_btn(),width=20)
		self.Button_admin_up.place(x=600, y=120)
		if(self.adminLogin==0):
			self.Button_admin_up.config(state='disable',relief=SUNKEN)

		###---------Refill BUTTON
		self.Button_view = tk.Button(self.GUIFrame_page_two, font=self.helv36_login, bd=1, text='VIEW STOCK ',
		                             command=lambda: self.viewStock_btn(), width=20)
		self.Button_view.place(x=600, y=200)

		###---------Refill BUTTON
		self.Button_ind = tk.Button(self.GUIFrame_page_two, font=self.helv36_login, bd=1, text='INDENT STOCK ',
		                             command=lambda: self.indentStock_btn(), width=20)
		self.Button_ind.place(x=600, y=280)

		###---------History BUTTON
		self.Button_htry = tk.Button(self.GUIFrame_page_two, font=self.helv36_login, bd=1, text='HISTORY ',
		                             command=lambda: self.history_btn(), width=20)
		self.Button_htry.place(x=600, y=360)

		###---------Logout BUTTON
		self.Button_log_out = tk.Button(self.GUIFrame_page_two, font=self.helv36_login, bd=1, text='LOGOUT ',
									 command=lambda: self.logout_btn(), width=15)
		self.Button_log_out.place(x=100, y=500)

	def logout_btn(self):
		self.GUIFrame_page_two.destroy()
		self.page_one()

#####################################ADMIN BUTTON########################################################
	def admin_update_btn(self):
		print("admin button")
		self.GUIFrame_page_two.destroy()
		self.GUIFrame_admin_update = tk.Frame(self.GUIFrame_default, width=900, height=600)
		self.GUIFrame_admin_update.pack(expand=False, anchor=CENTER)
		self.GUIFrame_admin_update.place(x=0, y=0)

		self.ltext2 = tk.Label(self.GUIFrame_admin_update, image=self.photo_westtek)
		self.ltext2.place(x=10, y=10)

		###---------Admin Logout BUTTON
		self.Button_log_out = tk.Button(self.GUIFrame_admin_update, font=self.helv36_login, bd=1, text='LOGOUT ',
										command=lambda: self.logoutAdmin_btn(), width=10)
		self.Button_log_out.place(x=100, y=500)

		###---------Admin back BUTTON
		self.Button_backAdmin = tk.Button(self.GUIFrame_admin_update, font=self.helv36_login, bd=1, text='BACK ',
										command=lambda: self.adminBack_btn(), width=10)
		self.Button_backAdmin.place(x=600, y=500)

		###---------Admin  Material code
		self.adminMaterialCode_Label = tk.Label(self.GUIFrame_admin_update, font=self.helv36_login, text="Material Code")
		self.adminMaterialCode_Label.place(x=50, y=150)

		self.adminMaterialCode_Var = StringVar()
		self.adminMaterialCode_Entry = tk.Entry(self.GUIFrame_admin_update, font=self.helv36_login, textvariable=self.adminMaterialCode_Var,
									  width=15)
		self.adminMaterialCode_Entry.place(x=50, y=180)

		###---------Admin  Material Description
		self.adminMaterialDesc_Label = tk.Label(self.GUIFrame_admin_update, font=self.helv36_login, text="Material Description")
		self.adminMaterialDesc_Label.place(x=270, y=150)

		self.adminMaterialDesc_Var = StringVar()
		self.adminMaterialDesc_Entry = tk.Entry(self.GUIFrame_admin_update, font=self.helv36_login,
											textvariable=self.adminMaterialDesc_Var,width=25)
		self.adminMaterialDesc_Entry.place(x=270, y=180)

		###---------Admin  store location
		self.adminStoreLoc_Label = tk.Label(self.GUIFrame_admin_update, font=self.helv36_login, text="Location")
		self.adminStoreLoc_Label.place(x=600, y=150)

		self.adminStoreLoc_Var = StringVar()
		self.adminStoreLoc_Entry = tk.Entry(self.GUIFrame_admin_update, font=self.helv36_login,
											textvariable=self.adminStoreLoc_Var,width=10)
		self.adminStoreLoc_Entry.place(x=600, y=180)

		###---------Admin  Batch
		self.adminBatch_Label = tk.Label(self.GUIFrame_admin_update, font=self.helv36_login, text="Batch")
		self.adminBatch_Label.place(x=50, y=240)

		self.adminBatch_Var = StringVar()
		self.adminBatch_Entry = tk.Entry(self.GUIFrame_admin_update, font=self.helv36_login,
											textvariable=self.adminBatch_Var, width=10)
		self.adminBatch_Entry.place(x=50, y=280)

		###---------Admin  Unrestricted
		self.adminUnrestricted_Label = tk.Label(self.GUIFrame_admin_update, font=self.helv36_login, text="Unrestricted")
		self.adminUnrestricted_Label.place(x=250, y=240)

		self.adminUnrestricted_Var = StringVar()
		self.adminUnrestricted_Entry = tk.Entry(self.GUIFrame_admin_update, font=self.helv36_login,
											textvariable=self.adminUnrestricted_Var, width=10)
		self.adminUnrestricted_Entry.place(x=250, y=280)

		###---------Admin  Measure
		self.adminMeasure_Label = tk.Label(self.GUIFrame_admin_update, font=self.helv36_login, text="Measure")
		self.adminMeasure_Label.place(x=450, y=240)

		self.adminMeasure_Var = StringVar()
		self.adminMeasure_Entry = tk.Entry(self.GUIFrame_admin_update, font=self.helv36_login,
											textvariable=self.adminMeasure_Var, width=10)
		self.adminMeasure_Entry.place(x=450, y=280)

		##-Admin Update button
		self.adminUpdateBtn = tk.Button(self.GUIFrame_admin_update, font=self.helv36_login, bd=2, text=' UPDATE ',
									command=lambda: self.admin_update_value(),image=self.update_log,compound=LEFT)
		self.adminUpdateBtn.place(x=550, y=400)

	def adminBack_btn(self):
		self.GUIFrame_admin_update.destroy()
		self.page_two()

	def logoutAdmin_btn(self):
		self.GUIFrame_admin_update.destroy()
		self.page_one()

	def admin_update_value(self):
		wb = openpyxl.load_workbook("/home/king/PycharmProjects/project_stores/data_storage/test_data.xlsx")
		sheet = wb["Sheet1"]
		row = sheet.max_row + 1
		update_value = [sheet.max_row ,self.adminMaterialCode_Var.get(), self.adminMaterialDesc_Var.get(),
						self.adminStoreLoc_Var.get(), self.adminBatch_Var.get(), self.adminUnrestricted_Var.get(),
						self.adminMeasure_Var.get()]
		print(update_value)
		for col, entry in enumerate(update_value, start=1):
			sheet.cell(row=row, column=col, value=entry)

		wb.save("/home/king/PycharmProjects/project_stores/data_storage/test_data.xlsx")
#########################################VIEW BUTTON#################################################
	def viewStock_btn(self):
		print("view stock")
		self.GUIFrame_page_two.destroy()
		self.GUIFrame_view_stock = tk.Frame(self.GUIFrame_default, width=900, height=600)
		self.GUIFrame_view_stock.pack(expand=False, anchor=CENTER)
		self.GUIFrame_view_stock.place(x=0, y=0)

		self.ltext2 = tk.Label(self.GUIFrame_view_stock, image=self.photo_westtek)
		self.ltext2.place(x=10, y=10)

		# ---scrollbar frame
		self.GUIFramescrollbar = tk.Frame(self.GUIFrame_view_stock, width=600, height=400)
		self.GUIFramescrollbar.place(x=20, y=100)

		# ---scrollbar initilization
		self.scrollbar_y = Scrollbar(self.GUIFramescrollbar, width=40)
		self.scrollbar_y.pack(side=RIGHT, fill=Y)

		self.scrollbar_x = Scrollbar(self.GUIFramescrollbar, width=32, orient=HORIZONTAL)
		self.scrollbar_x.pack(side=BOTTOM, fill=X)

		# ---textbox frame
		self.txt = tk.Text(self.GUIFramescrollbar, height=20, width=100, wrap=tk.NONE,
						   xscrollcommand=self.scrollbar_x.set,
						   yscrollcommand=self.scrollbar_y.set)
		self.txt.pack()

		# ---adding scrollbar to text
		self.scrollbar_y.config(command=self.txt.yview)
		self.scrollbar_x.config(command=self.txt.xview)

		self.txt.configure(state='normal')
		self.txt.delete(1.0, END)
		for i in range(1, self.row_count + 1):
			for j in range(1, self.column_count + 1):
				data = self.sheet.cell(row=i, column=j).value
				if (j == 1):
					self.txt.insert(tk.END, str(data).ljust(5)+' | ')
				elif (j == 2):
					self.txt.insert(tk.END, str(data).ljust(10)+' | ')
				elif(j == 3):
					self.txt.insert(tk.END, str(data).ljust(40)+' | ')
				elif(j == 4):
				 	self.txt.insert(tk.END, str(data).ljust(18)+' | ')
				elif (j == 5):
					self.txt.insert(tk.END, str(data).ljust(10)+' | ')
				elif (j == 6):
					self.txt.insert(tk.END, str(data).ljust(12)+' | ')
				elif (j == 7):
					self.txt.insert(tk.END, str(data).ljust(10)+' | ')
			self.txt.insert(tk.INSERT,'\n')
		self.txt.configure(state='disable')
		###--------- view Logout BUTTON
		self.Button_log_out = tk.Button(self.GUIFrame_view_stock, font=self.helv36_login, bd=1, text='LOGOUT ',
										command=lambda: self.logoutview_btn(), width=10)
		self.Button_log_out.place(x=50, y=500)

		###---------view back BUTTON
		self.Button_backAdmin = tk.Button(self.GUIFrame_view_stock, font=self.helv36_login, bd=1, text='BACK ',
										command=lambda: self.viewBack_btn(), width=10)
		self.Button_backAdmin.place(x=650, y=500)

	def viewBack_btn(self):
		self.GUIFrame_view_stock.destroy()
		self.page_two()

	def logoutview_btn(self):
		self.GUIFrame_view_stock.destroy()
		self.page_one()

	#########################################INDENT BUTTON#################################################
	def indentStock_btn(self):
		print("indent stock")
		self.GUIFrame_page_two.destroy()
		self.GUIFrame_indentStock = tk.Frame(self.GUIFrame_default, width=900, height=600)
		self.GUIFrame_indentStock.pack(expand=False, anchor=CENTER)
		self.GUIFrame_indentStock.place(x=0, y=0)

		self.ltext2 = tk.Label(self.GUIFrame_indentStock, image=self.photo_westtek)
		self.ltext2.place(x=10, y=10)

		if(self.loginName=="westtek"):
			#####-----read user name
			wb_obj_userName = openpyxl.load_workbook(
				"/home/king/PycharmProjects/project_stores/data_storage/user_data.xlsx")
			sheet_obj_userName = wb_obj_userName.active
			m_row_userName = sheet_obj_userName.max_row
			self.options = []
			for i in range(1, m_row_userName + 1):
				if (i > 1):
					cell_obj1 = sheet_obj_userName.cell(row=i, column=5)
					self.options.append(cell_obj1.value)

			###### datatype of menu text
			self.clicked = StringVar()

			###### initial menu text
			self.clicked.set("Select User")

			#####Create Dropdown menu
			self.drop = tk.OptionMenu(self.GUIFrame_indentStock, self.clicked, *self.options,command = self.OptionMenu_CheckButton)
			self.drop.config(width=12, font=self.helv36_login)
			menu = self.GUIFrame_indentStock.nametowidget(self.drop.menuname)
			menu.config(font=self.helv36_login)  # Set the dropdown menu's font
			self.drop.place(x=650, y=100)

			###---------Accept button
			self.Button_acceptIndentAdmin = tk.Button(self.GUIFrame_indentStock, font=self.helv36_login, bd=1,
											 text='Accept  ',
											 command=lambda: self.indentAcceptAdmin_btn(), width=10)
			self.Button_acceptIndentAdmin.place(x=450, y=150)

			###---------Reject button
			self.Button_rejectIndentAdmin = tk.Button(self.GUIFrame_indentStock, font=self.helv36_login, bd=1,
											 text='Reject ',
											 command=lambda: self.indentRejectAdmin_btn(), width=10)
			self.Button_rejectIndentAdmin.place(x=450, y=200)

			###---------clear button
			self.Button_ClearIndentAdmin = tk.Button(self.GUIFrame_indentStock, font=self.helv36_login, bd=1, text='Clear  ',
											command=lambda: self.clearUpdateAdmin_btn(), width=10)
			self.Button_ClearIndentAdmin.place(x=450, y=250)

			###---------Update button
			self.Button_updateIndentAdmin = tk.Button(self.GUIFrame_indentStock, font=self.helv36_login, bd=1,
											 text='Update ',
											 command=lambda: self.indentUpdateAdmin_btn(), width=10)
			self.Button_updateIndentAdmin.place(x=450, y=300)

########### text Box with message
			self.userIndent = tk.Text(self.GUIFrame_indentStock, height=15, width=35, bg='white', state=DISABLED,selectborderwidth=3)
			self.userIndent.place(x=620, y=200)
			self.userIndent.bind('<Triple-1>', self.userIndent_button)
			self.userIndent_lab = tk.Label(self.GUIFrame_indentStock, text="Indent files", compound="top"
									 , font=self.times_16)
			self.userIndent_lab.place(x=700, y=160)

		else:
			#####-----Select item label
			self.indentItems_Label = tk.Label(self.GUIFrame_indentStock, font=self.helv36_login, text="Select Items")
			self.indentItems_Label.place(x=450, y=150)

			#####-----read store details for custom combo box
			wb_obj_indent = openpyxl.load_workbook(
				"/home/king/PycharmProjects/project_stores/data_storage/store_details.xlsx")

			sheet_obj_indent = wb_obj_indent.active
			m_row_indent = sheet_obj_indent.max_row
			test_list = []
			for i in range(1, m_row_indent + 1):
				cell_obj1 = sheet_obj_indent.cell(row=i, column=3)
				test_list.append(cell_obj1.value)

			#####-----costom combo box
			self.box_value = tk.StringVar()
			self.combo = tkentrycomplete.AutocompleteCombobox(self.GUIFrame_indentStock, textvariable=self.box_value)
			self.combo.config(font=self.times_16)
			self.combo.config(width=20)
			self.combo.set_completion_list(test_list)
			self.combo.place(x=450, y=200)

			###---------material quantity label
			self.indentQty_Label = tk.Label(self.GUIFrame_indentStock, font=self.helv36_login, text="Qty")
			self.indentQty_Label.place(x=750, y=150)

			###---------material quantity entry
			self.indentQty_Var = StringVar()
			self.indentQty_Entry = tk.Entry(self.GUIFrame_indentStock, font=self.helv36_login,
											textvariable=self.indentQty_Var, width=10)
			self.indentQty_Entry.place(x=750, y=200)

			###---------Add button
			self.Button_addIndent = tk.Button(self.GUIFrame_indentStock, font=self.helv36_login, bd=1,
												 text='Add    ',
												 command=lambda: self.indentAdd_btn(), width=10)
			self.Button_addIndent.place(x=500, y=300)

			###---------Remove button
			self.Button_removeIndent = tk.Button(self.GUIFrame_indentStock, font=self.helv36_login, bd=1,
												 text='Remove ',
												 command=lambda: self.indentRemove_btn(), width=10)
			self.Button_removeIndent.place(x=500, y=350)

			###---------clear button
			self.Button_ClearIndent = tk.Button(self.GUIFrame_indentStock, font=self.helv36_login, bd=1, text='Clear  ',
												command=lambda: self.clearUpdate_btn(), width=10)
			self.Button_ClearIndent.place(x=700, y=300)

			###---------Update button
			self.Button_updateIndent = tk.Button(self.GUIFrame_indentStock, font=self.helv36_login, bd=1,
												 text='Update ',
												 command=lambda: self.indentUpdate_btn(), width=10)
			self.Button_updateIndent.place(x=700, y=350)

		###---------Admin  Measure
		self.indentQty_Label = tk.Label(self.GUIFrame_indentStock, font=self.times_16 , text="Qty")
		self.indentQty_Label.place(x=50, y=120)

		self.indentMaterial_Label = tk.Label(self.GUIFrame_indentStock, font=self.times_16 , text="Material Name")
		self.indentMaterial_Label.place(x=150, y=120)

		# ---scrollbar frame
		self.GUIFramescrollbarIndent = tk.Frame(self.GUIFrame_indentStock, width=400, height=400)
		self.GUIFramescrollbarIndent.place(x=30, y=150)

		# ---scrollbar initilization
		self.scrollbar_yIndent = Scrollbar(self.GUIFramescrollbarIndent, width=40)
		self.scrollbar_yIndent.pack(side=RIGHT, fill=Y)

		self.scrollbar_xIndent = Scrollbar(self.GUIFramescrollbarIndent, width=32, orient=HORIZONTAL)
		self.scrollbar_xIndent.pack(side=BOTTOM, fill=X)

		# ---textbox frame
		self.txtIndent = tk.Text(self.GUIFramescrollbarIndent, height=20, width=50, wrap=tk.NONE,
						   xscrollcommand=self.scrollbar_xIndent.set,
						   yscrollcommand=self.scrollbar_yIndent.set)
		self.txtIndent.pack()
		self.txtIndent.bind('<Triple-1>', self.Indent_text_button)
		# ---adding scrollbar to text
		self.scrollbar_yIndent.config(command=self.txtIndent.yview)
		self.scrollbar_xIndent.config(command=self.txtIndent.xview)

		self.txtIndent.configure(state='normal')
		self.txtIndent.delete(1.0, END)

###--------- indent Logout BUTTON
		self.Button_log_out = tk.Button(self.GUIFrame_indentStock, font=self.helv36_login, bd=1, text='LOGOUT ',
										command=lambda: self.logoutindent_btn(), width=10)
		self.Button_log_out.place(x=100, y=500)

###---------indent back BUTTON
		self.Button_backAdmin = tk.Button(self.GUIFrame_indentStock, font=self.helv36_login, bd=1, text='BACK ',
										  command=lambda: self.indentBack_btn(), width=10)
		self.Button_backAdmin.place(x=600, y=500)

	def indentBack_btn(self):
		self.Button_backAdmin.destroy()
		self.page_two()

	def logoutindent_btn(self):
		self.Button_backAdmin.destroy()
		self.page_one()

	def indentRemove_btn(self):
		self.txtIndent.configure(state='normal')
		self.txtIndent.delete(self.start_d, self.end_d)
		self.txtIndent.configure(state='disable')
		self.line=0

	def indentAdd_btn(self):
		if(self.box_value.get()==""):
			tk.messagebox.showerror("Error", "Material Name")
		elif(self.indentQty_Var.get()==""):
			tk.messagebox.showerror("Error", "Quantity")
		else:
			self.txtIndent.configure(state='normal')
			self.txtIndent.mark_set("insert", END)
			self.txtIndent.insert(tk.END, str(self.indentQty_Var.get()).ljust(5)+' | ')
			self.txtIndent.insert(tk.END, str(self.box_value.get()))
			self.txtIndent.insert(tk.INSERT,'\n')
			self.txtIndent.configure(state='disable')

###---text triple click event
	def Indent_text_button(self, event):
		index = self.txtIndent.index("@%s,%s" % (event.x, event.y))
		self.line, char = index.split(".")
		self.start_d = self.line + '.0'
		self.end_d = str(int(self.line) + 1) + '.0'
		if(self.loginName=="westtek"):
			pass
			print("westtek admin")

####---text clear event
	def clearUpdate_btn(self):
		self.txtIndent.configure(state='normal')
		self.txtIndent.delete(1.0, END)
		self.txtIndent.configure(state='disable')

	####----Update text data to excel with naming data and time
	def indentUpdate_btn(self):
		wbIndent = openpyxl.Workbook()
		sheet = wbIndent.active
		text=self.txtIndent.get("1.0",END).splitlines()
		sheet.cell(row=1, column=1).value = "Sl No"
		sheet.cell(row=1, column=2).value = "Material Description"
		sheet.cell(row=1, column=3).value = "Quantity"

		data_len=len(text)-1
		for i in range(data_len):
			x = text[i].split("|")
			sheet.cell(row=i + 2, column=1).value = i+1
			sheet.cell(row=i + 2, column=2).value = x[1]
			sheet.cell(row=i + 2, column=3).value =x[0].strip()
		ts = time.time()
		st = datetime.datetime.fromtimestamp(ts).strftime('%Y%m%d%H%M%S')
		wbIndent.save("//home/king/Desktop/"+self.loginName+st+".xlsx")
		self.txtIndent.delete(1.0, END)



######---Function related to Indent admin
	def indentAcceptAdmin_btn(self):
		pass

	def indentRejectAdmin_btn(self):
		pass

	def clearUpdateAdmin_btn(self):
		self.userIndent.configure(state='normal')
		self.userIndent.delete(1.0, END)
		self.userIndent.configure(state='disable')
		pass

	def indentUpdateAdmin_btn(self):
		pass

	def OptionMenu_CheckButton(self,event):
		self.entries = os.listdir('/home/king/PycharmProjects/project_stores/data_storage/user_indent/'+self.clicked.get()+'/')
		#self.entries=self.entries1.sort()
		self.entries= sorted(self.entries,reverse=True)
		self.userIndent.configure(state='normal')
		self.userIndent.delete(1.0, END)
		for entry in self.entries:
			self.userIndent.insert(tk.END, str(entry))
			self.userIndent.insert(tk.INSERT, '\n')
		self.userIndent.configure(state='disable')
		
	def userIndent_button(self,event):
		index = self.userIndent.index("@%s,%s" % (event.x, event.y))
		line, char = index.split(".")
		if(len(self.entries)>=int(line)):
			print(self.entries[int(line)-1])
			self.userIndent.configure(state='normal')
			self.userIndent.delete(1.0, END)
			self.wb_IndentIndvid = openpyxl.load_workbook("/home/king/PycharmProjects/project_stores/data_storage/user_indent/Surendar_t/Surendar_t20210617215735.xlsx")
			self.sheet_IndentIndvid = self.wb_IndentIndvid["Sheet"]
			self.row_count_IndentIndvid = self.sheet_IndentIndvid.max_row
			self.column_count_IndentIndvid = self.sheet_IndentIndvid.max_column

			for i in range(1, self.row_count_IndentIndvid + 1):
				for j in range(1, self.column_count_IndentIndvid + 1):
					data = self.sheet_IndentIndvid.cell(row=i, column=j).value
					if (j == 1):
						self.txtIndent.insert(tk.END, str(data).ljust(5) + ' | ')
					elif (j == 2):
						self.txtIndent.insert(tk.END, str(data).ljust(50) + ' | ')
					elif (j == 3):
						self.txtIndent.insert(tk.END, str(data).ljust(10) + ' | ')
				self.txtIndent.insert(tk.INSERT, '\n')
			self.txtIndent.configure(state='disable')


#########################################HISTORY BUTTON#################################################
	def history_btn(self):
		print("history")
		self.GUIFrame_page_two.destroy()
		self.GUIFrame_history = tk.Frame(self.GUIFrame_default, width=900, height=600)
		self.GUIFrame_history.pack(expand=False, anchor=CENTER)
		self.GUIFrame_history.place(x=0, y=0)

		###---------Logout BUTTON
		self.Button_history_out = tk.Button(self.GUIFrame_history, font=self.helv36_login, bd=1, text='LOGOUT ',
										command=lambda: self.logouthistory_btn(), width=15)
		self.Button_history_out.place(x=100, y=500)

		###---------indent back BUTTON
		self.Button_backHistory = tk.Button(self.GUIFrame_history, font=self.helv36_login, bd=1, text='BACK ',
										  command=lambda: self.historyBack_btn(), width=10)
		self.Button_backHistory.place(x=600, y=500)

		self.ltext2 = tk.Label(self.GUIFrame_history, image=self.photo_westtek)
		self.ltext2.place(x=10, y=10)

		# ---scrollbar frame
		self.GUIFramescrollbarHistory = tk.Frame(self.GUIFrame_history, width=400, height=600)
		self.GUIFramescrollbarHistory.place(x=30, y=110)

		# ---scrollbar initilization
		self.scrollbar_yIndent = Scrollbar(self.GUIFramescrollbarHistory, width=40)
		self.scrollbar_yIndent.pack(side=RIGHT, fill=Y)

		self.scrollbar_xIndent = Scrollbar(self.GUIFramescrollbarHistory, width=32, orient=HORIZONTAL)
		self.scrollbar_xIndent.pack(side=BOTTOM, fill=X)

		# ---textbox frame
		self.txtIndent_history = tk.Text(self.GUIFramescrollbarHistory, height=22, width=110, wrap=tk.NONE,
								 xscrollcommand=self.scrollbar_xIndent.set,
								 yscrollcommand=self.scrollbar_yIndent.set)
		self.txtIndent_history.pack()
		self.txtIndent_history.bind('<Triple-1>', self.on_text_button)

		# ---adding scrollbar to text
		self.scrollbar_yIndent.config(command=self.txtIndent_history.yview)
		self.scrollbar_xIndent.config(command=self.txtIndent_history.xview)

		self.txtIndent_history.configure(state='normal')
		self.txtIndent_history.delete(1.0, END)

		self.wb_history = openpyxl.load_workbook(str(data_folder) + "/data_storage/history_details.xlsx")
		self.sheet_history = self.wb_history["Sheet1"]
		self.row_count_history = self.sheet_history.max_row
		self.column_count_history = self.sheet_history.max_column

		for i in range(1, self.row_count_history + 1):
			for j in range(1, self.column_count_history + 1):
				data = self.sheet_history.cell(row=i, column=j).value
				if (j == 1):
					self.txtIndent_history.insert(tk.END, str(data).ljust(5) + ' | ')
				elif (j == 2):
					self.txtIndent_history.insert(tk.END, str(data).ljust(20) + ' | ')
				elif (j == 3):
					self.txtIndent_history.insert(tk.END, str(data).ljust(20) + ' | ')
				elif (j == 4):
					self.txtIndent_history.insert(tk.END, str(data).ljust(20) + ' | ')
				elif (j == 5):
					self.txtIndent_history.insert(tk.END, str(data).ljust(20) + ' | ')

			self.txtIndent_history.insert(tk.INSERT, '\n')
		self.txtIndent_history.configure(state='disable')

	def historyBack_btn(self):
		self.GUIFrame_history.destroy()
		self.page_two()

	def logouthistory_btn(self):
		self.GUIFrame_history.destroy()
		self.page_one()

	def on_text_button(self, event):
		index = self.txtIndent_history.index("@%s,%s" % (event.x, event.y))
		line, char = index.split(".")

rpi_UI()