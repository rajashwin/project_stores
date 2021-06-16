from openpyxl import load_workbook
from pathlib import Path
def xlread():
    data_folder = "/home/king/PycharmProjects/project_stores"
    file = "home/king/PycharmProjects/project_stores/data_storage/store_details.xlsx"
    new_row = ['data5', 'data6', 'data7', 'data8']

    wb = load_workbook("/home/king/PycharmProjects/project_stores/data_storage/test_data.xlsx")
    sheet = wb["Sheet1"]
    row = sheet.max_row+1
    for col, entry in enumerate(new_row, start=1):
        sheet.cell(row=row, column=col, value=entry)

    wb.save("/home/king/PycharmProjects/project_stores/data_storage/test_data.xlsx")






# workbook object is created
wb_obj = load_workbook("/home/king/PycharmProjects/project_stores/data_storage/user_data.xlsx")

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
username="Chethan.a"
password="Chet@ach"
# Loop will print all values
# of first column
for i in range(1, m_row + 1):
    cell_obj1 = sheet_obj.cell(row=i, column=5)
    cell_obj2 = sheet_obj.cell(row=i, column=6)
    if(username==cell_obj1.value and password==cell_obj2.value):
        print(cell_obj1.value,cell_obj2.value)
