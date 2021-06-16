import time
from tkinter import *
import tkinter as tk
from tkinter import font
import tkinter.messagebox
import tkentrycomplete

from openpyxl import load_workbook

root = tk.Tk()
box_value = tk.StringVar()
times_16 = font.Font(family="Times", size=16, weight='normal')
# workbook object is created
wb_obj = load_workbook("/home/king/PycharmProjects/project_stores/data_storage/store_details.xlsx")

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
test_list = []
def fun():
    print(box_value.get())


for i in range(1, m_row + 1):
    cell_obj1 = sheet_obj.cell(row=i, column=3)
    print(cell_obj1.value)
    test_list.append(cell_obj1.value)

combo = tkentrycomplete.AutocompleteCombobox(textvariable=box_value)

combo.set_completion_list(test_list)
combo.place(x=140, y=50)
combo.config(font = times_16)
button = tk.Button(text='but', command=fun)
button.place(x=140,y=70)

root.mainloop()